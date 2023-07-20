from Menu import Menu
import openpyxl
import pandas as pd


def clear_sheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None


def clear():
    filename = "filtro.xlsx"
    workbook = openpyxl.load_workbook(filename)

    # Selecione a planilha desejada
    sheet = workbook['Sheet1']

    # Limpe todos os dados da planilha
    clear_sheet(sheet)

    # Salve as alterações no arquivo
    workbook.save(filename)
def extrair_nomes(filename):
    names = []

    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook['Sheet1']

        for row in sheet.iter_rows(values_only=True):
            name = row[1]  # Assuming the "nome" column is the first column (index 0)
            names.append(name)

    except FileNotFoundError:
        print(f"O arquivo '{filename}' não foi encontrado.")

    names.pop(0)
    return names

def extrair_cursos(curso):
    # Ler o arquivo "cadeiras.xlsx" na sheet "main"
    df = pd.read_excel('cadeiras.xlsx', sheet_name='main')

    # Filtrar as linhas que correspondem ao curso
    filtro = df['curso'] == curso
    cursos_encontrados = df[filtro]

    # Verificar se foram encontrados cursos
    if cursos_encontrados.empty:
        return

    # Criar um novo DataFrame apenas com as linhas encontradas
    novo_df = pd.DataFrame(cursos_encontrados)

    # Salvar o novo DataFrame em um arquivo xlsx
    clear()
    novo_df.to_excel('filtro.xlsx', index=False)




def extrair_ano(ano):
    # Ler o arquivo "cadeiras.xlsx" na sheet "main"
    df = pd.read_excel('filtro.xlsx', sheet_name='Sheet1')

    # Filtrar as linhas que correspondem ao curso
    filtro = df['ano'] == ano
    cadeiras = df[filtro]

    # Verificar se foram encontrados cursos
    if cadeiras.empty:
        return

    # Criar um novo DataFrame apenas com as linhas encontradas
    novo_df = pd.DataFrame(cadeiras)

    # Salvar o novo DataFrame em um arquivo xlsx
    novo_df.to_excel('filtro.xlsx', index=False)


def extrair_semestre(semestre):
    # Ler o arquivo "cadeiras.xlsx" na sheet "main"
    df = pd.read_excel('filtro.xlsx', sheet_name='Sheet1')

    # Filtrar as linhas que correspondem ao curso
    filtro = df['semestre'] == semestre
    cadeiras = df[filtro]

    # Verificar se foram encontrados cursos
    if cadeiras.empty:
        return

    # Criar um novo DataFrame apenas com as linhas encontradas
    novo_df = pd.DataFrame(cadeiras)

    # Salvar o novo DataFrame em um arquivo xlsx
    novo_df.to_excel('filtro.xlsx', index=False)



account_sid = 'ACa5c83e45677a24409033440f8499727b'
auth_token = 'c1b4fb1c0d36280276b9f8a0b6a68cf2'
invalid_comand_error = 'Comando não reconhecido\nFormato de comando:\n    [curso]/[cadeira]/[nr de estudante]'
username = 'yannick.matimbe'
password = 'Y@nnick2003'
login_url = 'https://fenix.isutc.ac.mz/cas/login?service=https://fenix.isutc.ac.mz/isutc/'
logged_url = 'https://fenix.isutc.ac.mz/isutc/fenixEduIndex.do'
mark_url = 'https://fenix.isutc.ac.mz/isutc/publico/executionCourse.do?method=marks&executionCourseID=1126312223704031'
cursos = ["lca", "lea", "lect", "lecc", "lemec", "leet", "lee", "lef", "leit", "lemt", "lgbs", "lgf"]
invalid_semester_error = 'Semestre inválido.'

# extrair_cursos('leit')
# extrair_ano('1º Ano')
# extrair_semestre('1')

main_menu = Menu('main_menu', 'Selecione a operação: ', ['Cursos', 'Sobre'])
menu_sobre = Menu('menu_sobre', 'Selecione o opção: ', [])
menu_cursos = Menu('menu_cursos', 'Selecione um curso: ', cursos)
menu_ano = Menu('menu_ano', 'Selecione o ano acadêmico: ', ['1º Ano', '2º Ano', '3º Ano', '4º Ano'])
menu_semestre = Menu('menu_semestre', 'Selecione o semestre: ', ['1', '2'])
menu_cadeiras = Menu('menu_cadeira', 'Selecione a cadeira', extrair_nomes('filtro.xlsx'))
status = 'null'

print(main_menu.print_prompt())





